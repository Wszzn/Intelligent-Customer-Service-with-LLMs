from retriever import Retriever
from kg_builder import KGBuilder
from reranker import Reranker
import json
from embedding_handler import EmbeddingHandler
import time
from typing import List, Dict
from openai import OpenAI
from LLM_tool import LLMHandler
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
import os

class ChatSystem:
    def __init__(self):
        """初始化时加载所有预处理结果"""
        # 初始化嵌入处理器并加载索引
        time_start = time.time()
        self.embedding_handler = EmbeddingHandler(
            model_type="qwen", 
            load_from="/root/ty/data/embeddings"
        )
        self.embedding_handler.load_index("/root/ty/data/embeddings")
        print(self.embedding_handler.index is not None)
        time_end = time.time()
        print(f"加载索引耗时：{time_end - time_start}s")
        # 加载分块数据
        # with open("/root/ty/data/chunks.json", "r", encoding="utf-8") as f:
        #     self.chunks = json.load(f)
        
        # 初始化KG构建器并加载索引
        self.kg_builder = KGBuilder(
            embeddingHandler=self.embedding_handler
        )
        self.kg_builder.load_vector_indexes("/root/ty/data/vector_indexes")
        # print(self.kg_builder.entity_index is not None)
        # 初始化检索器
        self.retriever = Retriever(self.embedding_handler, self.kg_builder)
        self.reranker = Reranker()
        self.llm_handler = LLMHandler()
        self.client = OpenAI(api_key="your_api_key", base_url="https://api.deepseek.com")
    def process_query(self, query: str) -> list:
        """处理用户查询的完整流程"""
        # 0. 预处理查询文本
        query = self._preprocess_query(query)

        # 1. 检索实体和关系
        entity_chunk_ids = []
        relation_chunk_ids = []
        entity_kg_results = []
        ralation_kg_results = []
        keywords = self.llm_handler.keywords_extraction(query)
        entity_kg_results,ralation_kg_results = self.kg_builder.query_knowledge_graph(keywords["low_level_keywords"],keywords["high_level_keywords"])
        seen = set()
        entity_kg_results = [
            x for x in entity_kg_results 
            if tuple(sorted(x['entity'].items())) not in seen 
            and not seen.add(tuple(sorted(x['entity'].items())))
        ]
        
        # 同样方式处理relation_kg_results
        seen = set()
        ralation_kg_results = [
            x for x in ralation_kg_results 
            if tuple(sorted(x['relation'].items())) not in seen 
            and not seen.add(tuple(sorted(x['relation'].items())))
        ]
        # entity_kg_results = list(set(entity_kg_results))
        # ralation_kg_results = list(set(ralation_kg_results))
        # print(f"从知识图谱中检索出的结果：{kg_results}\n{'--'*50}")
        # print(f"从知识图谱中检索出的结果：{entity_kg_results}\n关系：{ralation_kg_results}\n{'--'*50}")
        entity_chunk_ids = [entity_kg['entity']["chunkid"] for entity_kg in entity_kg_results]
        # entity_chunk_ids = list(set(entity_chunk_ids))
        relation_chunk_ids = [relation_kg['relation']["chunkid"] for relation_kg in ralation_kg_results]
        # relation_chunk_ids = list(set(relation_chunk_ids))
        # entity_chunk = [self.embedding_handler.documents[chunkid] for chunkid in entity_chunk_ids]
        # relation_chunk = [self.embedding_handler.documents[chunkid] for chunkid in relation_chunk_ids]
        # print(f'检索出的实体{entity_kg_results},检索出来的关系{ralation_kg_results}\n{"++"*50}')
        # 2. 检索相关文档
        retrieved_results = self.retriever.search(query)
        # print(f"文档检索出的结果：{retrieved_results}\n{'--'*50}")
        retrieved_chunk_ids = [result['chunkid'] for result in retrieved_results]
        chunk_ids = list(set(entity_chunk_ids + relation_chunk_ids + retrieved_chunk_ids))
        # chunk_ids = list(set(chunk_ids))
        retrieved_chunk = [self.embedding_handler.documents[chunkid] for chunkid in chunk_ids]
        # print(f"文档检索出的结果：{retrieved_chunk[:10]}\n{'--'*50}")
        # reranked_retrieved_results = []

        # 3. 重新排序

        reranked_retrieved_results = self.reranker.rerank(retrieved_chunk, query)

        # print(f"检索出的结果：{reranked_retrieved_results[:10]}\n{'--'*50}")
        return reranked_retrieved_results,entity_kg_results,ralation_kg_results

    def _preprocess_query(self, query: str) -> str:
        """预处理查询文本"""
        import re
        # 去除前后空白
        query = query.strip()
        # 去除多余空格
        query = re.sub(r'\s+', ' ', query)
        # 其他可能的预处理步骤...
        return query

    def ask(self, query: str) -> dict:
        """
        完整问答流程：处理查询并生成最终回答
        
        Args:
            query: 用户问题
            
        Returns:
            包含检索结果和生成回答的字典
        """
        # 处理查询获取相关结果
        reranked_retrieved_results, entity_kg_results, ralation_kg_results = self.process_query(query)
        # print("实体是非空的++++++++++++++++++++++++++++++++++++",entity_kg_results is not None)
        if entity_kg_results:
            # print("调用混合环境","*"*100)
            # 将实体结果转换为Markdown表格
            entity_table = "| 赛事 | 实体名称 | 类型 | 描述 |\n|----------|----------|------|------|\n"
            for entity in entity_kg_results:
                entity_table += f"| {self.embedding_handler.metadata[entity['entity']['chunkid']]} | {entity['entity'].get('name', '')} | {entity['entity'].get('type', '')} | {entity['entity'].get('description', '')} |\n"
            
            # 将关系结果转换为Markdown表格
            relation_table = "| 赛事 | 主体 | 关系 | 客体 | 权重 | 描述 |\n|------|------|------|------|------|------|\n"
            for relation in ralation_kg_results:
                relation_table += f"| {self.embedding_handler.metadata[relation['relation']['chunkid']]} | {relation['relation'].get('subject', '')} | {relation['relation'].get('predicate', '')} | {relation['relation'].get('object', '')} | {relation['relation'].get('strength', '')} | {relation['relation'].get('description', '')} |\n"
            
            kg_data = entity_table + "\n\n" + relation_table
            # print(f"混合环境下的上下文：{kg_data}\n{'++'*50}")
            answer = self.llm_handler.mix_rag_response(query, reranked_retrieved_results[:5], kg_data)
            # print(f"混合环境下的回答：{answer}\n{'++'*50}")
            return {
                "question": answer["question"],
                "answer": answer["answer"],
                "keyword": answer["keyword"],
                "references": reranked_retrieved_results[:5],  # 返回前5个参考结果
                "kg_data": kg_data
            }
            
        else:
            print("调用本地环境","*"*100)
            
            # 生成最终回答
            # answer = self.llm_handler.naive_rag_response(query, reranked_retrieved_results)
            answer = self.generate_final_answer(query, reranked_retrieved_results[:5])
            return {
                "answer": answer,
                "references": reranked_retrieved_results[:5]  # 返回前5个参考结果
            }

    def generate_final_answer(self, query: str, context_results: List[Dict]) -> str:
        """
        使用大模型根据检索结果生成最终回答
        
        Args:
            query: 用户查询
            context_results: 检索到的上下文结果列表
            
        Returns:
            生成的最终回答
        """
        # 构建上下文提示
        context_prompt = "根据以下信息回答问题：\n"
        for i, result in enumerate(context_results[:5], 1):  # 最多使用前5个结果
            context_prompt += f"{i}. {result.get('description', result.get('document', '无内容'))}\n"
        
        # 构建完整提示
        prompt = f"""你是一个专业的比赛赛事问答助手，请根据提供的上下文信息，用简洁准确的语言回答用户的问题。
        
        上下文信息：
        {context_prompt}

        用户问题：{query}
        若上下文没有提供相关信息，请明确告知无法回答，不能胡编乱造。
        语言要简介，但是信息一定要完整。
        请直接给出答案，不要包含无关的解释或说明。"""
        
        # 调用大模型生成回答
        response = self.client.chat.completions.create(
            model="deepseek-chat",
            messages=[
                {"role": "system", "content": prompt},
                {"role": "user", "content": query},
            ],
            temperature=0.3,
            stream=False
        )
        
        return response.choices[0].message.content

def save_qa_to_excel(queries, answers, output_file="result_2.xlsx"):
    """将问答结果保存到Excel文件"""
    # 创建数据框
    data = []
    for i, (query, answer) in enumerate(zip(queries, answers), 1):
        # 生成4位数的编号
        qa_id = f"C{str(i).zfill(4)}"
        
        # 提取关键点 (假设关键点在answer字典中)
        # key_points = answer.get("key_points", "")  # 如果没有关键点，使用空字符串
        
        data.append({
            "问题编号": qa_id,
            "问题": query,
            "关键点": answer["keyword"],
            "回答": answer["answer"]
        })
    
    # 创建DataFrame
    df = pd.DataFrame(data)
    
    # 创建Excel文件
    wb = Workbook()
    ws = wb.active
    
    # 写入表头
    headers = ["问题编号", "问题", "关键点", "回答"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    for row, item in enumerate(data, 2):
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = item[key]
            cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # 调整列宽
    ws.column_dimensions['A'].width = 12  # 问题编号
    ws.column_dimensions['B'].width = 40  # 问题
    ws.column_dimensions['C'].width = 30  # 关键点
    ws.column_dimensions['D'].width = 60  # 回答
    
    # 保存文件
    wb.save(output_file)
    print(f"问答结果已保存到: {output_file}")

if __name__ == "__main__":
    """命令行交互入口"""
    system = ChatSystem()
    queries = [
        "第七届全国青少年人工智能创新挑战赛的报名时间是什么时候？",
        "“未来校园智能应用专项赛”中智能交通信号灯任务的基本要求是什么？",
        "3D编程模型创新设计专项赛中“伞”设计需要考虑哪些方面？",
        "“人工智能综合创新专项赛”参赛作品的字数要求是什么？",
        "“未来校园智能应用专项赛”中哪些任务涉及到自动化控制？",
        "如何确保我的竞赛作品不被剽窃？",
        "“未来校园智能应用专项赛”参赛前有哪些准备工作？",
        "在3D编程模型创新设计专项赛中，提交的任务分数不显示或出现错误，怎么办？",
        "第七届全国青少年人工智能创新挑战赛中有多少个机器人类别的竞赛？",
        "我是一名学生家长，孩子现在是高中一年级，他的编程能力很强，但是动手制作能力相对较弱，我想问一下，参加“第七届全国青少年人工智能创新挑战赛”中哪一项（或哪一类）的竞赛比较合适。",
        # "人工智能相关竞赛有多少？",
        # "如何准备第七届全国青少年人工智能创新挑战赛"
    ]

    # for query in queries:
    #     results = system.process_query(query)
    #     print(f"查询：{query}")
    #     print("\n查询结果:")
    #     for i, result in enumerate(results, 1):  # 显示前5个结果
    #         print(f"{i}. {result.get('document', result.get('name', '无标题'))}")
    #     print("\n\n")

    #     print("知识问答系统已启动，输入'退出'结束对话")
    # while True:
    #     query = input("请输入您的问题: ")
    #     if query.lower() == "退出":
    #         break
            
    #     results = system.process_query(query)
    #     print("\n查询结果:")
    #     for i, result in enumerate(results[:5], 1):  # 显示前5个结果
    #         print(f"{i}. {result.get('document', result.get('name', '无标题'))}")



    # answers = []  # 存储所有回答
    # for query in queries:
    #     results = system.ask(query)
    #     print(f"查询：{query}")
    #     print("查询结果:", results["answer"])
    #     print("-"*50)
    #     answers.append(results)  # 保存回答结果
    #     print("\n\n")
    
    # # 保存到Excel文件
    # save_qa_to_excel(queries, answers)
    
    print("知识问答系统已启动，输入'退出'结束对话")
    while True:
        query = input("请输入您的问题: ")
        if query.lower() == "退出":
            break
            
        results = system.ask(query)
        print("\n查询的关键点",results["keyword"])
        print("查询结果:",results["answer"])
        print("-"*50)
        
